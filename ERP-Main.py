import sys
import mysql.connector
import PySimpleGUI as sg
import openpyxl
from openpyxl import load_workbook
import pyodbc
import datetime
import os
from decimal import Decimal
from cryptography.fernet import Fernet
import appdirs
import base64

key = b'X='
cipher_suite = Fernet(key)

appname = "SPP"
appauthor = "Estaq"
datapath = appdirs.user_data_dir(appname, appauthor)
os.makedirs(datapath, exist_ok=True)  # Ensure the directory exists

login_file_path = os.path.join(datapath, "login.txt")

# Connect to the MySQL database
mydb = mysql.connector.connect(
    host="estaq.eng.br",
    user="X",
    password="X",
    database="X",
    port=X
)

sg.set_options(background_color='#10b010')
sg.set_options(text_element_background_color='#10b010')
sg.set_options(button_color=('text_color', 'button_color'))
sg.set_options(button_color=('white', 'black'))

appdata_path = os.getenv('APPDATA')

icon_path = os.path.join(appdata_path, 'Local/Estaq/Spp/favicon.ico')

sg.set_options(icon=icon_path)

def verify_login(username, password):
    cursor = mydb.cursor()
    query = """
        SELECT *
        FROM usuarios
        WHERE login = %s AND senha = %s
    """
    cursor.execute(query, (username, password,))
    result = cursor.fetchone()
    return result is not None


def show_login_window():
    try:
        with open(login_file_path, "rb") as f:
            encrypted_login = f.read()
        decrypted_login = cipher_suite.decrypt(
            encrypted_login).decode().split('\n')
        default_username = decrypted_login[0]
        default_password = decrypted_login[1]
        remember_me = True

        # Check if the credentials are valid
        if verify_login(default_username, default_password):
            show_main_window()  # Show the main window after a successful login
            return
    except Exception:
        default_username = ""
        default_password = ""
        remember_me = False

    layout = [
        [sg.Text("Usuario:", size=(7, 1)), sg.Input(
            default_username, key="-USERNAME-", size=(30, 1))],
        [sg.Text("Senha:", size=(7, 1)), sg.Input(default_password,
                                                  key="-PASSWORD-", password_char='*', size=(30, 1))],
        [sg.Column([[sg.Checkbox("Lembrar Login", default=remember_me, key="-REMEMBER-",
                   background_color='#10b010')]], background_color='#10b010')],
        [sg.Button("Login")]
    ]

    login_window = sg.Window("Login", layout)
    while True:
        event, values = login_window.read()
        if event == sg.WIN_CLOSED:
            login_window.close()
        elif event == "Login":
            username = values["-USERNAME-"]
            password = values["-PASSWORD-"]
            remember_me = values["-REMEMBER-"]
            if verify_login(username, password):
                if remember_me:
                    # Save the encrypted username and password to the file if "Remember me" is checked
                    login_info = f"{username}\n{password}"
                    encrypted_login = cipher_suite.encrypt(login_info.encode())
                    with open(login_file_path, "wb") as f:
                        f.write(encrypted_login)
                else:
                    # Delete the file if remember me is not checked
                    try:
                        os.remove(login_file_path)
                    except Exception:
                        pass
                login_window.close()
                show_main_window()  # Show the service window after a successful login
            else:
                sg.popup(
                    "Nome de usuário ou senha inválidos. Por favor, tente novamente.")


def get_proposta_codigo_options(selected_service):
    query1 = """
        SELECT DISTINCT pro.propostacodigo, BINARY emp.endereco, BINARY emp.cidade, BINARY emp.uf
        FROM propostas pro
        JOIN obras ob ON pro.idproposta = ob.idproposta
        JOIN empreendimentos emp ON pro.idempreendimento = emp.idempreendimento
        WHERE pro.propostasigla = %s AND ob.status_obra = '2.EM ANDAMENTO'
    """
    cursor = mydb.cursor()
    cursor.execute(query1, (selected_service,))

    # Get the results and decode binary data
    options = []
    for row in cursor.fetchall():
        options.append((row[0], f"{row[1].decode('utf-8') if isinstance(row[1], bytearray) else row[1]}, {row[2].decode('utf-8') if isinstance(row[2], bytearray) else row[2]}, {row[3].decode('utf-8') if isinstance(row[3], bytearray) else row[3]}"))

    return options


def get_idobra_for_proposta_codigo(proposta_codigo):
    query2 = """
        SELECT obras.idobra
        FROM propostas
        JOIN obras ON propostas.idproposta = obras.idproposta
        WHERE propostas.propostacodigo = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query2, (proposta_codigo,))
    result = cursor.fetchone()
    if result:
        return result[0]
    else:
        return None


def on_submit_proposta_hours(selected_value):
    proposta_codigo = selected_value.split(" - ")[0]
    cursor = mydb.cursor()
    query3 = """
        SELECT patrimonios.nome
        FROM propostas
        JOIN obras ON propostas.idproposta = obras.idproposta
        JOIN obras_patrimonios ON obras.idobra = obras_patrimonios.idobra
        JOIN patrimonios ON obras_patrimonios.idpatrimonio = patrimonios.idpatrimonio
        WHERE propostas.propostacodigo = %s
    """
    cursor.execute(query3, (proposta_codigo,))
    result = cursor.fetchall()

    layout = [
        [sg.Text(f"Patrimonios para a proposta {proposta_codigo}:")],
        [sg.Listbox(values=result, size=(40, 10), key="-LISTBOX-")],
        [sg.Button("OK"), sg.Button("Voltar")]
    ]
    results_window = sg.Window("Resultados", layout)

    while True:
        event, values = results_window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "OK":
            selected_patrimonio = values["-LISTBOX-"][0]
            results_window.close()
            show_hours_input_window_hours(proposta_codigo, selected_patrimonio)
            break
        elif event == "Voltar":
            results_window.close()
            break


def show_hours_input_window_hours(proposta_codigo, selected_patrimonio):
    cursor = mydb.cursor()
    query_patrimonio_id = """
        SELECT idpatrimonio
        FROM patrimonios
        WHERE nome = %s
    """
    cursor.execute(query_patrimonio_id, (selected_patrimonio[0],))
    idpatrimonio = cursor.fetchone()[0]

    query_descricao = """
        SELECT descricao
        FROM producaoocorrencias
    """
    cursor.execute(query_descricao)

    descricao_values = [row[0] for row in cursor.fetchall()]
    hours = [str(x).zfill(2) for x in range(0, 24)]
    minutes = [str(x).zfill(2) for x in range(0, 60)]

    layout = [
        [sg.Text(f"Patrimônio: {selected_patrimonio}")],
        [sg.Text("Ocorrência:"), sg.Listbox(
            values=descricao_values, size=(40, 5), key="-DESCRICAO-")],
        [sg.Text("Data de Lançamento:"), sg.Input(key="-DATE-", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="-DATE-", format="%Y-%m-%d")],
        [sg.Text("Hora Inicio:"), sg.Input("00:00", key="-START_TIME-", size=(5, 1), enable_events=True,
                                           text_color="black", background_color="white", do_not_clear=True, justification="center")],
        [sg.Text("Hora Final:"), sg.Input("00:00", key="-END_TIME-", size=(5, 1), enable_events=True,
                                          text_color="black", background_color="white", do_not_clear=True, justification="center")],
        [sg.Text("Qtd. Horas de Atraso:"), sg.Text(
            "", key="-TOTAL_HOURS-", size=(10, 1))],
        [sg.Text("Observações Complementares:"), sg.Input(
            key="-OBSERVACOES-", do_not_clear=True)],
        [sg.Button("Enviar"), sg.Button("Voltar")]
    ]

    hours_input_window = sg.Window("Horas", layout)

    def update_total_hours(window, start_hour, start_minute, end_hour, end_minute):
        if not start_hour or not start_minute or not end_hour or not end_minute:
            return 

        start_time = int(start_hour) * 60 + int(start_minute)
        end_time = int(end_hour) * 60 + int(end_minute)
        total_minutes = end_time - start_time
        if total_minutes < 0:
            return
        else:
            total_hours = total_minutes // 60
            remaining_minutes = total_minutes % 60
            window["-TOTAL_HOURS-"].update(
                f"{str(total_hours).zfill(2)}:{str(remaining_minutes).zfill(2)}")

    while True:
        event, values = hours_input_window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "-START_TIME-" or event == "-END_TIME-":
            if ':' in values["-START_TIME-"] and ':' in values["-END_TIME-"]:
                start_hour, start_minute = values["-START_TIME-"].split(':')
                end_hour, end_minute = values["-END_TIME-"].split(':')
                update_total_hours(hours_input_window, start_hour,
                                   start_minute, end_hour, end_minute)
        elif event == "Voltar":
            hours_input_window.close()
            on_submit_proposta_hours(proposta_codigo)
            break
        elif event == "Enviar":
            descricao = values["-DESCRICAO-"][0]
            date = values["-DATE-"]
            start_hour, start_minute = values["-START_TIME-"].split(':')
            end_hour, end_minute = values["-END_TIME-"].split(':')
            observacoes = values["-OBSERVACOES-"]
            end_time_minutes = int(end_hour) * 60 + int(end_minute)
            start_time_minutes = int(start_hour) * 60 + int(start_minute)
            time_difference = end_time_minutes - start_time_minutes
            total_hours = f"{str(time_difference // 60).zfill(2)}:{str(time_difference % 60).zfill(2)}"
            hours_input_window.close()

            cursor.execute(
                "SELECT idocorrencia FROM producaoocorrencias WHERE descricao = %s", (descricao,))
            idocorrencia = cursor.fetchone()[0]

            cursor.execute(
                "SELECT idobra FROM obras WHERE idproposta = (SELECT idproposta FROM propostas WHERE propostacodigo = %s)", (proposta_codigo,))
            idobra = cursor.fetchone()[0]

            cursor.execute(
                "SELECT idpatrimonio FROM patrimonios WHERE nome = %s", (selected_patrimonio[0],))

            idpatrimonio = cursor.fetchone()[0]

            query_insert = """
                INSERT INTO obras_ocorrencias (idocorrencia, idobra, data_lcto, horas_total, descricao, hora_inicial, hora_final, idpatrimonio)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query_insert, (idocorrencia, idobra, date, total_hours, observacoes,
                           f"{start_hour}:{start_minute}", f"{end_hour}:{end_minute}", idpatrimonio))
            mydb.commit()

            sg.popup("Dados inseridos com sucesso.", title="Sucesso")

            hours_input_window.close()
            show_hours_input_window_hours(proposta_codigo, selected_patrimonio)


def get_all_funcionarios():
    cursor = mydb.cursor()
    cursor.execute("SELECT idfuncionario, nome FROM producaofuncionarios")
    return cursor.fetchall()

def get_funcionario_options(idpatrimonio):
    cursor = mydb.cursor()
    query = "SELECT idfuncionario, nome FROM producaofuncionarios WHERE idpatrimonio = %s"
    cursor.execute(query, (idpatrimonio,))
    options = cursor.fetchall()
    cursor.close()
    return options

def show_update_funcionarios_window(proposta_codigo, selected_patrimonio):
    cursor1 = mydb.cursor()
    query_patrimonio_id = """
        SELECT patrimonios.idpatrimonio
        FROM patrimonios
        WHERE patrimonios.nome = %s
    """
    cursor1.execute(query_patrimonio_id, (selected_patrimonio[0],))
    idpatrimonio, = cursor1.fetchone()
    cursor1.fetchall()
    cursor1.close()

    cursor2 = mydb.cursor()
    cursor2.execute("SELECT idfuncionario, nome FROM producaofuncionarios")
    all_funcionario_data = cursor2.fetchall()
    all_funcionario_values = [
        f"{row[0]} - {row[1]}" for row in all_funcionario_data]
    cursor2.fetchall()
    cursor2.close()

    cursor3 = mydb.cursor()
    cursor3.execute(
        "SELECT idfuncionario, nome FROM producaofuncionarios WHERE idpatrimonio = %s", (idpatrimonio,))
    current_funcionario_data = cursor3.fetchall()
    current_funcionario_values = [
        f"{row[0]} - {row[1]}" for row in current_funcionario_data]
    cursor3.fetchall()
    cursor3.close()

    layout = [
        [sg.Text(f"Patrimônio: {selected_patrimonio}")],
        *[[sg.Combo(all_funcionario_values, default_value=current_funcionario_values[i] if i < len(
            current_funcionario_values) else "", key=f"FUNCIONARIO_{i}", enable_events=True)] for i in range(8)],
        [sg.Button("Atualizar"), sg.Button("Voltar")]
    ]

    update_funcionarios_window = sg.Window("Atualizar Funcionarios", layout)

    while True:
        event, values = update_funcionarios_window.read()
        if event in (sg.WIN_CLOSED, "Voltar"):
            update_funcionarios_window.close()
            break
        elif event == "Atualizar":
            update_funcionarios_window.close()

            cursor3 = mydb.cursor()
            cursor3.execute(
                "UPDATE producaofuncionarios SET idpatrimonio = 0 WHERE idpatrimonio = %s", (idpatrimonio,))
            cursor3.fetchall()
            cursor3.close()

            cursor4 = mydb.cursor()
            for i in range(8):
                if f"FUNCIONARIO_{i}" in values and values[f"FUNCIONARIO_{i}"]:
                    idfuncionario = int(
                        values[f"FUNCIONARIO_{i}"].split(" - ")[0])
                    cursor4.execute(
                        "UPDATE producaofuncionarios SET idpatrimonio = %s WHERE idfuncionario = %s", (idpatrimonio, idfuncionario))
            cursor4.fetchall()
            cursor4.close()

            show_hours_input_window_teams(proposta_codigo, selected_patrimonio)
            break


def on_submit_proposta_teams(selected_value):
    proposta_codigo = selected_value.split(" - ")[0]

    cursor = mydb.cursor()
    query3 = """
        SELECT patrimonios.nome
        FROM propostas
        JOIN obras ON propostas.idproposta = obras.idproposta
        JOIN obras_patrimonios ON obras.idobra = obras_patrimonios.idobra
        JOIN patrimonios ON obras_patrimonios.idpatrimonio = patrimonios.idpatrimonio
        WHERE propostas.propostacodigo = %s
    """
    cursor.execute(query3, (proposta_codigo,))
    result = cursor.fetchall()

    layout = [
        [sg.Text(f"Patrimonios para a proposta {proposta_codigo}:")],
        [sg.Listbox(values=result, size=(40, 10), key="-LISTBOX-")],
        [sg.Button("OK"), sg.Button("Voltar")]
    ]
    results_window = sg.Window("Resultados", layout)

    while True:
        event, values = results_window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "OK":
            selected_patrimonio = values["-LISTBOX-"][0]
            results_window.close()
            show_hours_input_window_teams(proposta_codigo, selected_patrimonio)
            break
        elif event == "Voltar":
            results_window.close()
            break


def register_funcionario_window():
    layout = [
        [sg.Text("Nome:"), sg.Input(key="-NOME-")],
        [sg.Button("Confirmar")]
    ]
    register_window = sg.Window("Registrar Funcionario", layout)

    while True:
        event, values = register_window.read()
        if event == sg.WIN_CLOSED:
            register_window.close()
            break
        elif event == "Confirmar":
            nome = values["-NOME-"]
            if nome == "":
                sg.popup("Por favor, insira um nome.")
            else:
                nome = nome.upper()  # convert input to uppercase
                cursor = mydb.cursor()
                cursor.execute(
                    "INSERT INTO producaofuncionarios (nome, funcao, equipe, idpatrimonio) VALUES (%s, %s, %s, 0)", (nome, '', ''))
                mydb.commit()
                cursor.close()
                register_window.close()
                break


def show_hours_input_window_teams(proposta_codigo, selected_patrimonio):
    cursor1 = mydb.cursor()
    query_patrimonio_id = """
        SELECT patrimonios.idpatrimonio, obras.idobra
        FROM patrimonios
        JOIN obras_patrimonios ON patrimonios.idpatrimonio = obras_patrimonios.idpatrimonio
        JOIN obras ON obras_patrimonios.idobra = obras.idobra
        JOIN propostas ON propostas.idproposta = obras.idproposta
        WHERE patrimonios.nome = %s AND propostas.propostacodigo = %s
    """
    cursor1.execute(query_patrimonio_id,
                    (selected_patrimonio[0], proposta_codigo))
    idpatrimonio, idobra = cursor1.fetchone()
    cursor1.fetchall()
    cursor1.close()

    cursor2 = mydb.cursor()
    cursor2.execute(
        "SELECT idfuncionario, nome FROM producaofuncionarios WHERE idpatrimonio = %s", (idpatrimonio,))
    funcionario_data = cursor2.fetchall()
    print(funcionario_data)
    funcionario_options = get_funcionario_options(idpatrimonio)
    funcionario_values = [
        f"{str(option[0])} - {option[1]}" for option in funcionario_options]
    cursor2.fetchall()
    cursor2.close()

    ocorrencia = ""
    equipe = ""
    data_entrada1 = "07:30:00"
    data_saida1 = "12:00:00"
    data_entrada2 = "13:00:00"
    data_saida2 = "17:30:00"

    funcionario_lines = []
    for i in range(8):
        if i < len(funcionario_values):
            funcionario_lines.append([sg.Text(funcionario_values[i])])
    layout = [
        [sg.Text(f"Patrimônio: {selected_patrimonio}")],
        *funcionario_lines,
        [sg.Text("Data de Lançamento:"), sg.Input(key="-DATE-", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="-DATE-", format="%Y-%m-%d")],
        [sg.Button("Enviar"), sg.Button("Atualizar"), sg.Button(
            "Registrar"), sg.Button("Voltar")]
    ]

    hours_input_window = sg.Window("Equipes", layout)

    while True:
        event, values = hours_input_window.read()
        if event == sg.WIN_CLOSED:
            hours_input_window.close()
            break
        elif event == "Voltar":
            hours_input_window.close()
            on_submit_proposta_teams(proposta_codigo)
            break
        elif event == "Registrar":
            hours_input_window.close()
            register_funcionario_window()
            break
        elif event == "Atualizar":
            hours_input_window.close()
            show_update_funcionarios_window(
                proposta_codigo, selected_patrimonio)
            break
        elif event == "Enviar":
            data_lcto = values["-DATE-"]

            confirm = sg.popup_yes_no(
                "Você realmente deseja inserir os dados?")
            if confirm == "Yes":
                insert_query = """
                    INSERT INTO obras_funcionarios (idfuncionario, idobra, data_lcto, idpatrimonio, ocorrencia, equipe, data_entrada1, data_saida1, data_entrada2, data_saida2)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                insert_cursor = mydb.cursor()
                for row in funcionario_data:
                    idfuncionario = row[0]
                    # Debug
                    print(f"Inserting for idobra: {idobra}")
                    insert_cursor.execute(insert_query, (idfuncionario, idobra, data_lcto, idpatrimonio,
                                          ocorrencia, equipe, data_entrada1, data_saida1, data_entrada2, data_saida2))

                mydb.commit()
                sg.popup("Dados inseridos com sucesso.", title="Sucesso")


def on_submit_proposta(proposta_codigo, selected_service, selected_machine=None, file_path=None):
    idobra = get_idobra_for_proposta_codigo(proposta_codigo)

    cursor = mydb.cursor()
    query3 = """
        SELECT patrimonios.idpatrimonio, patrimonios.nome
        FROM propostas
        JOIN obras ON propostas.idproposta = obras.idproposta
        JOIN obras_patrimonios ON obras.idobra = obras_patrimonios.idobra
        JOIN patrimonios ON obras_patrimonios.idpatrimonio = patrimonios.idpatrimonio
        WHERE propostas.propostacodigo = %s
    """
    cursor.execute(query3, (proposta_codigo,))
    result = cursor.fetchall()

    formatted_result = [f"{row[0]} - {row[1]}" for row in result]

    if file_path:
        query = """
            SELECT propostas.idproposta
            FROM propostas
            WHERE propostas.propostacodigo = %s
        """
        cursor = mydb.cursor()
        cursor.execute(query, (proposta_codigo,))
        result = cursor.fetchone()
        idproposta = result[0]

        if selected_service == 'RT':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            valorfaturminimo = check_v_produzido_sum_rt(idproposta)
            process_rt_service(ws, selected_idpatrimonio, idobra,
                               selected_service, idproposta, valorfaturminimo)

        elif selected_service == 'RZ':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            valorfaturminimo = check_v_produzido_sum_rz(idproposta)
            process_rz_service(ws, selected_idpatrimonio, idobra,
                               selected_service, idproposta, valorfaturminimo)

        elif selected_service == 'MT':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            process_mt_service(ws, selected_idpatrimonio,
                               idobra, selected_service, idproposta)

        elif selected_service == 'PM':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            process_pm_service(ws, selected_idpatrimonio,
                               idobra, selected_service, idproposta)

        elif selected_service == 'PD':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            valorfaturminimo = check_v_produzido_sum_pd(idproposta)
            process_pd_service(ws, selected_idpatrimonio, idobra,
                               selected_service, idproposta, valorfaturminimo)

        elif selected_service == 'TR':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            process_tr_service(ws, selected_idpatrimonio,
                               idobra, selected_service, idproposta)

        elif selected_service == 'GR':
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            process_gr_service(ws, selected_idpatrimonio,
                               idobra, selected_service, idproposta)

        elif selected_service == 'HC':
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={file_path};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            mdb_columns = [
                'estNumero', 'estDiametro', 'estInicioP', 'estFimP',
                'estInicioC', 'estFimC', 'estComprimento', 'estVolBetao',
                'estSuperConsumo'
            ]
            mdb_query = f"SELECT {', '.join(mdb_columns)} FROM ESTACA"
            cursor.execute(mdb_query)

            selected_idpatrimonio = int(selected_machine.split(" - ")[0])
            valorfaturminimo = check_v_produzido_sum_hc(idproposta)
            process_hc_service(cursor, selected_idpatrimonio, idobra,
                               selected_service, idproposta, valorfaturminimo)

    else:
        layout = [
            [sg.Text(f"Patrimonios para a proposta {proposta_codigo}:")],
            [sg.Listbox(values=formatted_result,
                        size=(40, 10), key="-LISTBOX-")],
            [sg.Button("OK")]
        ]
        results_window = sg.Window("Resultados", layout)

        while True:
            event, values = results_window.read()
            if event == sg.WIN_CLOSED:
                break
            elif event == "OK":
                selected = values["-LISTBOX-"]
                if len(selected) > 0:
                    selected_machine = selected[0]
                    show_machine_window(
                        selected_machine, selected_service, proposta_codigo)
                break

        results_window.close()


def on_submit_proposta_manual(proposta_codigo, selected_service, selected_machine=None):
    idobra = get_idobra_for_proposta_codigo(proposta_codigo)

    cursor = mydb.cursor()
    query3 = """
        SELECT patrimonios.idpatrimonio, patrimonios.nome
        FROM propostas
        JOIN obras ON propostas.idproposta = obras.idproposta
        JOIN obras_patrimonios ON obras.idobra = obras_patrimonios.idobra
        JOIN patrimonios ON obras_patrimonios.idpatrimonio = patrimonios.idpatrimonio
        WHERE propostas.propostacodigo = %s
    """
    cursor.execute(query3, (proposta_codigo,))
    result = cursor.fetchall()

    formatted_result = [f"{row[0]} - {row[1]}" for row in result]

    query = """
        SELECT propostas.idproposta
        FROM propostas
        WHERE propostas.propostacodigo = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (proposta_codigo,))
    result = cursor.fetchone()
    idproposta = result[0]

    layout = [
        [sg.Text(f"Patrimonios para a proposta {proposta_codigo}:")],
        [sg.Listbox(values=formatted_result, size=(40, 10), key="-LISTBOX-")],
        [sg.Button("OK")]
    ]
    results_window = sg.Window("Resultados", layout)

    while True:
        event, values = results_window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "OK":
            selected = values["-LISTBOX-"]
            if len(selected) > 0:
                selected_machine = selected[0]

                if selected_service == 'RT':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    valorfaturminimo = check_v_produzido_sum_rt_manual(
                        idproposta)
                    process_rt_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo)

                elif selected_service == 'PD':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    valorfaturminimo = check_v_produzido_sum_pd_manual(
                        idproposta)
                    process_pd_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo)

                elif selected_service == 'RZ':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    valorfaturminimo = check_v_produzido_sum_rz_manual(
                        idproposta)
                    process_rz_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo)

                elif selected_service == 'MT':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    process_mt_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta)

                elif selected_service == 'PM':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    process_pm_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta)

                elif selected_service == 'TR':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    process_tr_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta)

                elif selected_service == 'GR':
                    selected_idpatrimonio = int(
                        selected_machine.split(" - ")[0])
                    process_gr_service_manual(
                        selected_idpatrimonio, idobra, selected_service, idproposta)

            break

    results_window.close()


def insert_rt_data_into_db(ordered_rt_data):
    cursor = mydb.cursor()

    seca_submersa = ordered_rt_data.pop('Seca / Submersa', None)
    ordered_rt_data['es_flag_seca'] = 'S' if seca_submersa == 'Seca' else 'N'
    ordered_rt_data['es_flag_submersa'] = 'S' if seca_submersa == 'Submersa' else 'N'

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, concretagem_prevista, concretagem_realizada, concretagem_altura, es_flag_seca, es_flag_submersa, diametroalargamento, profundidadealargamento, es_diametro, es_profundidade, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(concretagem_prevista)s, %(concretagem_realizada)s, %(concretagem_altura)s, %(es_flag_seca)s, %(es_flag_submersa)s, %(diametroalargamento)s, %(profundidadealargamento)s, %(es_diametro)s, %(es_profundidade)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_rt_data)
    mydb.commit()


def insert_rt_data_into_db_manual(rt_data):
    cursor = mydb.cursor()

    seca_submersa = rt_data.pop('Seca / Submersa', None)
    rt_data['es_flag_seca'] = 'S' if seca_submersa == 'Seca' else 'N'
    rt_data['es_flag_submersa'] = 'S' if seca_submersa == 'Submersa' else 'N'

    # Convert the values to .
    rt_data['diametro'] = float(rt_data['diametro'].replace(',', '.'))
    rt_data['secaoprojeto'] = float(rt_data['secaoprojeto'].replace(',', '.'))
    rt_data['profundidade'] = float(rt_data['profundidade'].replace(',', '.'))
    rt_data['concretagem_altura'] = float(
        rt_data['concretagem_altura'].replace(',', '.'))
    rt_data['diametroalargamento'] = float(
        rt_data['diametroalargamento'].replace(',', '.'))
    rt_data['profundidadealargamento'] = float(
        rt_data['profundidadealargamento'].replace(',', '.'))
    rt_data['es_diametro'] = float(rt_data['es_diametro'].replace(',', '.'))
    rt_data['es_profundidade'] = float(
        rt_data['es_profundidade'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, concretagem_prevista, concretagem_realizada, concretagem_altura, es_flag_seca, es_flag_submersa, diametroalargamento, profundidadealargamento, es_diametro, es_profundidade, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(concretagem_prevista)s, %(concretagem_realizada)s, %(concretagem_altura)s, %(es_flag_seca)s, %(es_flag_submersa)s, %(diametroalargamento)s, %(profundidadealargamento)s, %(es_diametro)s, %(es_profundidade)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(rt_data)

    cursor.execute(sql_query, rt_data)
    mydb.commit()


def insert_rz_data_into_db(ordered_rz_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, rz_solo, rz_alteracao, rz_rocha, profundidade, concretagem_inicio, concretagem_termino, rz_sacos_cimento, rz_areia, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(rz_solo)s, %(rz_alteracao)s, %(rz_rocha)s, %(profundidade)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(rz_sacos_cimento)s, %(rz_areia)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_rz_data)
    mydb.commit()


def insert_rz_data_into_db_manual(rz_data):
    cursor = mydb.cursor()

    # Convert the values to .
    rz_data['diametro'] = float(rz_data['diametro'].replace(',', '.'))
    rz_data['secaoprojeto'] = float(rz_data['secaoprojeto'].replace(',', '.'))
    rz_data['rz_solo'] = float(rz_data['rz_solo'].replace(',', '.'))
    rz_data['rz_alteracao'] = float(rz_data['rz_alteracao'].replace(',', '.'))
    rz_data['rz_rocha'] = float(rz_data['rz_rocha'].replace(',', '.'))
    rz_data['rz_sacos_cimento'] = float(
        rz_data['rz_sacos_cimento'].replace(',', '.'))
    rz_data['rz_areia'] = float(rz_data['rz_areia'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, rz_solo, rz_alteracao, rz_rocha, profundidade, concretagem_inicio, concretagem_termino, rz_sacos_cimento, rz_areia, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(rz_solo)s, %(rz_alteracao)s, %(rz_rocha)s, %(profundidade)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(rz_sacos_cimento)s, %(rz_areia)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(rz_data)

    cursor.execute(sql_query, rz_data)
    mydb.commit()


def insert_mt_data_into_db(ordered_mt_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, elementos, emendas, talas, cortes, nega, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(elementos)s, %(emendas)s, %(talas)s, %(cortes)s, %(nega)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_mt_data)
    mydb.commit()


def insert_mt_data_into_db_manual(mt_data):
    cursor = mydb.cursor()

    # Convert the values to .
    mt_data['diametro'] = float(mt_data['diametro'].replace(',', '.'))
    mt_data['secaoprojeto'] = float(mt_data['secaoprojeto'].replace(',', '.'))
    mt_data['profundidade'] = float(mt_data['profundidade'].replace(',', '.'))
    mt_data['elementos'] = float(mt_data['elementos'].replace(',', '.'))
    mt_data['emendas'] = float(mt_data['emendas'].replace(',', '.'))
    mt_data['talas'] = float(mt_data['talas'].replace(',', '.'))
    mt_data['cortes'] = float(mt_data['cortes'].replace(',', '.'))
    mt_data['nega'] = float(mt_data['nega'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, elementos, emendas, talas, cortes, nega, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(elementos)s, %(emendas)s, %(talas)s, %(cortes)s, %(nega)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(mt_data)

    cursor.execute(sql_query, mt_data)
    mydb.commit()


def insert_pm_data_into_db(ordered_pm_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, elementos, emendas, talas, cortes, nega, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(elementos)s, %(emendas)s, %(talas)s, %(cortes)s, %(nega)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_pm_data)
    mydb.commit()


def insert_pm_data_into_db_manual(pm_data):
    cursor = mydb.cursor()

    # Convert the values to .
    pm_data['diametro'] = float(pm_data['diametro'].replace(',', '.'))
    pm_data['secaoprojeto'] = float(pm_data['secaoprojeto'].replace(',', '.'))
    pm_data['profundidade'] = float(pm_data['profundidade'].replace(',', '.'))
    pm_data['elementos'] = float(pm_data['elementos'].replace(',', '.'))
    pm_data['emendas'] = float(pm_data['emendas'].replace(',', '.'))
    pm_data['talas'] = float(pm_data['talas'].replace(',', '.'))
    pm_data['cortes'] = float(pm_data['cortes'].replace(',', '.'))
    pm_data['nega'] = float(pm_data['nega'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, profundidade, elementos, emendas, talas, cortes, nega, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(profundidade)s, %(elementos)s, %(emendas)s, %(talas)s, %(cortes)s, %(nega)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(pm_data)

    cursor.execute(sql_query, pm_data)
    mydb.commit()


def insert_pd_data_into_db(ordered_pd_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, largura, area, concretagem_inicio, concretagem_termino, concretagem_realizada, concretagem_prevista, pd_flag_submersa, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(largura)s, %(area)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(concretagem_realizada)s, %(concretagem_prevista)s, %(pd_flag_submersa)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_pd_data)
    mydb.commit()


def insert_pd_data_into_db_manual(pd_data):

    # Convert the values to .
    pd_data['diametro'] = float(pd_data['diametro'].replace(',', '.'))
    pd_data['secaoprojeto'] = float(pd_data['secaoprojeto'].replace(',', '.'))
    pd_data['profundidade'] = float(pd_data['profundidade'].replace(',', '.'))
    pd_data['largura'] = float(pd_data['largura'].replace(',', '.'))
    pd_data['concretagem_realizada'] = float(
        pd_data['concretagem_realizada'].replace(',', '.'))

    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, largura, area, concretagem_inicio, concretagem_termino, concretagem_realizada, concretagem_prevista, pd_flag_submersa, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(largura)s, %(area)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(concretagem_realizada)s, %(concretagem_prevista)s, %(pd_flag_submersa)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(pd_data)

    cursor.execute(sql_query, pd_data)
    mydb.commit()


def insert_tr_data_into_db(ordered_tr_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, tr_datainjecao, tr_dataprojecao, rz_rocha, rz_solo, rz_sacos_cimento, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(tr_datainjecao)s, %(tr_dataprojecao)s, %(rz_rocha)s, %(rz_solo)s, %(rz_sacos_cimento)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_tr_data)
    mydb.commit()


def insert_tr_data_into_db_manual(tr_data):
    cursor = mydb.cursor()

    # Convert the values to .
    tr_data['diametro'] = float(tr_data['diametro'].replace(',', '.'))
    tr_data['secaoprojeto'] = float(tr_data['secaoprojeto'].replace(',', '.'))
    tr_data['profundidade'] = float(tr_data['profundidade'].replace(',', '.'))
    tr_data['rz_rocha'] = float(tr_data['rz_rocha'].replace(',', '.'))
    tr_data['rz_sacos_cimento'] = float(
        tr_data['rz_sacos_cimento'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, tr_datainjecao, tr_dataprojecao, rz_rocha, rz_solo, rz_sacos_cimento, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(tr_datainjecao)s, %(tr_dataprojecao)s, %(rz_rocha)s, %(rz_solo)s, %(rz_sacos_cimento)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(tr_data)

    cursor.execute(sql_query, tr_data)
    mydb.commit()


def insert_gr_data_into_db(ordered_gr_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, tr_datainjecao, tr_dataprojecao, rz_rocha, rz_solo, rz_sacos_cimento, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(tr_datainjecao)s, %(tr_dataprojecao)s, %(rz_rocha)s, %(rz_solo)s, %(rz_sacos_cimento)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_gr_data)
    mydb.commit()


def insert_gr_data_into_db_manual(gr_data):
    cursor = mydb.cursor()

    # Convert the values to .
    gr_data['diametro'] = float(gr_data['diametro'].replace(',', '.'))
    gr_data['secaoprojeto'] = float(gr_data['secaoprojeto'].replace(',', '.'))
    gr_data['profundidade'] = float(gr_data['profundidade'].replace(',', '.'))
    gr_data['rz_rocha'] = float(gr_data['rz_rocha'].replace(',', '.'))
    gr_data['rz_sacos_cimento'] = float(
        gr_data['rz_sacos_cimento'].replace(',', '.'))

    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, profundidade, tr_datainjecao, tr_dataprojecao, rz_rocha, rz_solo, rz_sacos_cimento, observacoes, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(tr_datainjecao)s, %(tr_dataprojecao)s, %(rz_rocha)s, %(rz_solo)s, %(rz_sacos_cimento)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    print(gr_data)

    cursor.execute(sql_query, gr_data)
    mydb.commit()


def insert_hc_data_into_db(ordered_hc_data):
    cursor = mydb.cursor()
    sql_query = f"""
    INSERT INTO obras_producao (idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto, hora_inicio, hora_termino, concretagem_inicio, concretagem_termino, profundidade, rz_solo, concretagem_realizada, concretagem_prevista, sobreconsumo, flag_aprovada, idservico, v_produzido)
    VALUES (%(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s, %(hora_inicio)s, %(hora_termino)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(profundidade)s, %(rz_solo)s, %(concretagem_realizada)s, %(concretagem_prevista)s, %(sobreconsumo)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s)
    """
    cursor.execute(sql_query, ordered_hc_data)
    mydb.commit()


def verify_and_insert_rt_data(ordered_rt_data, idproposta):
    v_produzido = 0
    sigla = ordered_rt_data["sigla"]
    diametro = float(ordered_rt_data["diametro"])

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_rt, obras.encamisamento, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_rt, obras.encamisamento, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if result[1] == diametro:  
            idservico = result[0]
            vlrunitario = result[2]
            encamisamento = result[4]
            profcobrancamin_rt = result[3]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = ordered_rt_data["profundidade"]
    diametro = ordered_rt_data["diametro"]
    seca_submersa = ordered_rt_data["Seca / Submersa"]
    concretagem_altura = ordered_rt_data["concretagem_altura"]

    concretagem_prevista = round(
        3.141592653589793 * ((diametro / 200) ** 2) * profundidade, 2)
    ordered_rt_data["concretagem_prevista"] = concretagem_prevista
    concretagem_realizada = round(
        3.141592653589793 * ((diametro / 200) ** 2) * concretagem_altura, 2)
    ordered_rt_data["concretagem_realizada"] = concretagem_realizada

    if seca_submersa == "Seca":
        es_flag_seca = "S"
        es_flag_submersa = "N"
    elif seca_submersa == "Submersa":
        es_flag_seca = "N"
        es_flag_submersa = "S"

    if idservico and vlrunitario is not None:
        ordered_rt_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db(ordered_rt_data)

        elif es_flag_seca == "N" and es_flag_submersa == "S":
            profundidade = max(float(profundidade), float(profcobrancamin_rt))
            encamisamento_corrected = float(encamisamento) / 100
            v_produzido = round(
                float(profundidade) * float(vlrunitario) * (1 + encamisamento_corrected), 2)
            ordered_rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db(ordered_rt_data)

        elif es_flag_seca == "S" and es_flag_submersa == "N":
            profundidade = max(float(profundidade), float(profcobrancamin_rt))
            v_produzido = round(float(profundidade) * float(vlrunitario))
            ordered_rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db(ordered_rt_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_rt_data_manual(rt_data, idproposta):
    v_produzido = 0
    sigla = rt_data["sigla"]
    diametro = float(rt_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_rt, obras.encamisamento, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_rt, obras.encamisamento, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if result[1] == diametro:
            idservico = result[0]
            vlrunitario = result[2]
            encamisamento = result[4]
            profcobrancamin_rt = result[3]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}, qtd {qtd}")

    profundidade = float(rt_data["profundidade"].replace(',', '.'))
    diametro = float(rt_data["diametro"].replace(',', '.'))
    concretagem_altura = float(rt_data["concretagem_altura"].replace(',', '.'))
    seca_submersa = rt_data["Seca / Submersa"]

    concretagem_prevista = round(
        3.141592653589793 * ((diametro / 200) ** 2) * profundidade, 2)
    rt_data["concretagem_prevista"] = concretagem_prevista
    concretagem_realizada = round(
        3.141592653589793 * ((diametro / 200) ** 2) * concretagem_altura, 2)
    rt_data["concretagem_realizada"] = concretagem_realizada

    if seca_submersa == "Seca":
        es_flag_seca = "S"
        es_flag_submersa = "N"
    elif seca_submersa == "Submersa":
        es_flag_seca = "N"
        es_flag_submersa = "S"

    if idservico and vlrunitario is not None:
        rt_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db_manual(rt_data)

        elif es_flag_seca == "N" and es_flag_submersa == "S":
            profundidade = max(float(profundidade), float(profcobrancamin_rt))
            encamisamento_corrected = float(encamisamento) / 100
            v_produzido = round(
                float(profundidade) * float(vlrunitario) * (1 + encamisamento_corrected), 2)
            rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db_manual(rt_data)

        elif es_flag_seca == "S" and es_flag_submersa == "N":
            profundidade = max(float(profundidade), float(profcobrancamin_rt))
            v_produzido = round(float(profundidade) * float(vlrunitario))
            rt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rt_data_into_db_manual(rt_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_rz_data(ordered_rz_data, idproposta):
    v_produzido = 0
    sigla = ordered_rz_data["sigla"]
    diametro = float(ordered_rz_data["diametro"])  # Convert diametro to float

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    vlrunitariorocha = None
    qtd = None
    for result in results:
        if result[1] == diametro:
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            valorpctefechado = result[4]
            break

    rz_solo = ordered_rz_data["rz_solo"]
    rz_alteracao = ordered_rz_data["rz_alteracao"]
    rz_rocha = ordered_rz_data["rz_rocha"]

    profundidade = float(rz_solo) + float(rz_alteracao) + float(rz_rocha)
    ordered_rz_data["profundidade"] = profundidade

    if idservico and vlrunitario is not None:
        ordered_rz_data["idservico"] = idservico

        if valorpctefechado > 0:
            print("Pacote Fechado Descoberto")
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_rz_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rz_data_into_db(ordered_rz_data)

        else:
            v_produzido1 = float(rz_solo) * float(vlrunitario)
            v_produzido2 = (float(rz_alteracao) + float(rz_rocha)
                            ) * float(vlrunitariorocha)
            v_produzido = round(v_produzido1 + v_produzido2, 2)
            ordered_rz_data["v_produzido"] = v_produzido
            insert_rz_data_into_db(ordered_rz_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_rz_data_manual(rz_data, idproposta):
    v_produzido = 0
    sigla = rz_data["sigla"]
    diametro = float(rz_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    vlrunitariorocha = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if result[1] == diametro: 
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            valorpctefechado = result[4]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    rz_solo = float(rz_data["rz_solo"].replace(',', '.'))
    rz_alteracao = float(rz_data["rz_alteracao"].replace(',', '.'))
    rz_rocha = float(rz_data["rz_rocha"].replace(',', '.'))

    profundidade = float(rz_solo) + float(rz_alteracao) + float(rz_rocha)
    rz_data["profundidade"] = profundidade

    if idservico and vlrunitario is not None:
        rz_data["idservico"] = idservico

        if valorpctefechado > 0:
            print("Pacote Fechado Descoberto")
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            rz_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_rz_data_into_db_manual(rz_data)

        else:
            v_produzido1 = float(rz_solo) * float(vlrunitario)
            v_produzido2 = (float(rz_alteracao) + float(rz_rocha)
                            ) * float(vlrunitariorocha)
            v_produzido = round(v_produzido1 + v_produzido2, 2)
            rz_data["v_produzido"] = v_produzido
            insert_rz_data_into_db_manual(rz_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        # Event loop for the warning window
        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_mt_data(ordered_mt_data, idproposta):
    v_produzido = 0
    sigla = ordered_mt_data["sigla"]
    diametro = float(ordered_mt_data["diametro"])  # Convert diametro to float

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_mt, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_mt, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if abs(float(result[1]) - diametro) < 0.001:
            idservico = result[0]
            vlrunitario = result[2]
            profcobrancamin_mt = result[3]
            valorpctefechado = result[4]
            break

    profundidade = ordered_mt_data["profundidade"]

    if idservico and vlrunitario is not None:
        ordered_mt_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_mt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db(ordered_mt_data)

        else:
            profundidade = max(float(profundidade), float(profcobrancamin_mt))
            v_produzido = round(float(profundidade) * float(vlrunitario), 2)
            ordered_mt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db(ordered_mt_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_mt_data_manual(mt_data, idproposta):
    v_produzido = 0
    sigla = mt_data["sigla"]
    diametro = float(mt_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_mt, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_mt, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if abs(float(result[1]) - diametro) < 0.001:
            idservico = result[0]
            vlrunitario = result[2]
            profcobrancamin_mt = result[3]
            valorpctefechado = result[4]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = float(mt_data["profundidade"].replace(',', '.'))

    if idservico and vlrunitario is not None:
        mt_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            mt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db_manual(mt_data)

        else:
            profundidade = max(float(profundidade), float(profcobrancamin_mt))
            v_produzido = round(float(profundidade) * float(vlrunitario), 2)
            mt_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db_manual(mt_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_pm_data(ordered_pm_data, idproposta):
    v_produzido = 0
    sigla = ordered_pm_data["sigla"]
    diametro = float(ordered_pm_data["diametro"])

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pm, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pm, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if abs(float(result[1]) - diametro) < 0.001:
            idservico = result[0]
            vlrunitario = result[2]
            profcobrancamin_pm = result[3]
            valorpctefechado = result[4]
            break

    profundidade = ordered_pm_data["profundidade"]

    if idservico and vlrunitario is not None:
        ordered_pm_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_pm_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db(ordered_pm_data)

        else:
            profundidade = max(float(profundidade), float(profcobrancamin_pm))
            v_produzido = round(float(profundidade) * float(vlrunitario), 2)
            ordered_pm_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_mt_data_into_db(ordered_pm_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_pm_data_manual(pm_data, idproposta):
    v_produzido = 0
    sigla = pm_data["sigla"]
    diametro = float(pm_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pm, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pm, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if abs(float(result[1]) - diametro) < 0.001:
            idservico = result[0]
            vlrunitario = result[2]
            profcobrancamin_pm = result[3]
            valorpctefechado = result[4]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = float(pm_data["profundidade"].replace(',', '.'))

    if idservico and vlrunitario is not None:
        pm_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            pm_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pm_data_into_db_manual(pm_data)

        else:
            profundidade = max(float(profundidade), float(profcobrancamin_pm))
            v_produzido = round(float(profundidade) * float(vlrunitario), 2)
            pm_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pm_data_into_db_manual(pm_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_pd_data(ordered_pd_data, idproposta):
    v_produzido = 0
    sigla = ordered_pd_data["sigla"]
    diametro = float(ordered_pd_data["diametro"])  # Convert diametro to float

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pd, obras.encamisamento, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pd, obras.encamisamento, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if result[1] == diametro: 
            idservico = result[0]
            vlrunitario = result[2]
            encamisamento = result[4]
            profcobrancamin_pd = result[3]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    pd_flag_submersa = "S"
    ordered_pd_data["pd_flag_submersa"] = pd_flag_submersa

    profundidade = ordered_pd_data["profundidade"]
    diametro = ordered_pd_data["diametro"]
    largura = ordered_pd_data["largura"]

    if idservico and vlrunitario is not None:
        ordered_pd_data["idservico"] = idservico
        area = round(float(profundidade) * float(largura), 2)
        ordered_pd_data["area"] = area
        concretagem_prevista = round(float(area) * float(diametro / 100), 2)
        ordered_pd_data["concretagem_prevista"] = concretagem_prevista

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_pd_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pd_data_into_db(ordered_pd_data)

        elif pd_flag_submersa == "S":
            profundidade = max(float(profundidade), float(profcobrancamin_pd))
            encamisamento_corrected = float(encamisamento) / 100
            print(float(encamisamento))
            v_produzido = round(float(area) * float(vlrunitario)
                                * (1 + encamisamento_corrected), 2)
            ordered_pd_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pd_data_into_db(ordered_pd_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_pd_data_manual(pd_data, idproposta):
    v_produzido = 0
    area = 0
    sigla = pd_data["sigla"]
    diametro = float(pd_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pd, obras.encamisamento, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_pd, obras.encamisamento, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if result[1] == diametro: 
            idservico = result[0]
            vlrunitario = result[2]
            encamisamento = result[4]
            profcobrancamin_pd = result[3]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    pd_flag_submersa = "S"
    pd_data["pd_flag_submersa"] = pd_flag_submersa

    profundidade = float(pd_data["profundidade"].replace(',', '.'))
    diametro = float(pd_data["diametro"].replace(',', '.'))
    largura = float(pd_data["largura"].replace(',', '.'))

    if idservico and vlrunitario is not None:
        pd_data["idservico"] = idservico
        area = round(float(profundidade) * float(largura), 2)
        pd_data["area"] = area
        concretagem_prevista = round(float(area) * float(diametro / 100), 2)
        pd_data["concretagem_prevista"] = concretagem_prevista

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            pd_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pd_data_into_db_manual(pd_data)

        elif pd_flag_submersa == "S":
            profundidade = max(profundidade, float(profcobrancamin_pd))
            encamisamento_corrected = float(encamisamento) / 100
            print(float(encamisamento))
            v_produzido = round(float(area) * float(vlrunitario)
                                * (1 + encamisamento_corrected), 2)
            pd_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_pd_data_into_db_manual(pd_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_tr_data(ordered_tr_data, idproposta):
    v_produzido = 0
    sigla = ordered_tr_data["sigla"]
    diametro = float(ordered_tr_data["diametro"])

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if float(result[1]) == diametro: 
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            profcobrancamin_tr = result[4]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = ordered_tr_data["profundidade"]
    rz_rocha = ordered_tr_data["rz_rocha"]

    rz_solo = round(float(profundidade) - float(rz_rocha), 2)
    ordered_tr_data["rz_solo"] = rz_solo

    if idservico and vlrunitario is not None:
        ordered_tr_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_tr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db(ordered_tr_data)

        else:
            profundidade_solo = max(float(rz_solo), float(
                profcobrancamin_tr) - float(rz_rocha))
            v_produzido = round(float(profundidade_solo) * float(vlrunitario) +
                                float(rz_rocha) * float(vlrunitariorocha), 2)
            ordered_tr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db(ordered_tr_data)

        return v_produzido


def verify_and_insert_tr_data_manual(tr_data, idproposta):
    v_produzido = 0
    sigla = tr_data["sigla"]
    diametro = float(tr_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if float(result[1]) == diametro: 
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            profcobrancamin_tr = result[4]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = float(tr_data["profundidade"].replace(',', '.'))
    rz_rocha = float(tr_data["rz_rocha"].replace(',', '.'))

    rz_solo = round(float(profundidade) - float(rz_rocha), 2)
    tr_data["rz_solo"] = rz_solo

    if idservico and vlrunitario is not None:
        tr_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            tr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db_manual(tr_data)

        else:
            profundidade_solo = max(float(rz_solo), float(
                profcobrancamin_tr) - float(rz_rocha))
            v_produzido = round(float(profundidade_solo) * float(vlrunitario) +
                                float(rz_rocha) * float(vlrunitariorocha), 2)
            tr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db_manual(tr_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_gr_data(ordered_gr_data, idproposta):
    v_produzido = 0
    sigla = ordered_gr_data["sigla"]
    diametro = float(ordered_gr_data["diametro"]) 

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if float(result[1]) == diametro:  
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            profcobrancamin_tr = result[4]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = ordered_gr_data["profundidade"]
    rz_rocha = ordered_gr_data["rz_rocha"]

    rz_solo = round(float(profundidade) - float(rz_rocha), 2)
    ordered_gr_data["rz_solo"] = rz_solo

    if idservico and vlrunitario is not None:
        ordered_gr_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_gr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db(ordered_gr_data)

        else:
            profundidade_solo = max(float(rz_solo), float(
                profcobrancamin_tr) - float(rz_rocha))
            v_produzido = round(float(profundidade_solo) * float(vlrunitario) +
                                float(rz_rocha) * float(vlrunitariorocha), 2)
            ordered_gr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_tr_data_into_db(ordered_gr_data)

        return v_produzido


def verify_and_insert_gr_data_manual(gr_data, idproposta):
    v_produzido = 0
    sigla = gr_data["sigla"]
    diametro = float(gr_data["diametro"].replace(',', '.'))

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, propostasitens.vlrunitariorocha, obras.profcobrancamin_tr, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if float(result[1]) == diametro:  
            idservico = result[0]
            vlrunitario = result[2]
            vlrunitariorocha = result[3]
            profcobrancamin_tr = result[4]
            valorpctefechado = result[5]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = float(gr_data["profundidade"].replace(',', '.'))
    rz_rocha = float(gr_data["rz_rocha"].replace(',', '.'))

    rz_solo = round(float(profundidade) - float(rz_rocha), 2)
    gr_data["rz_solo"] = rz_solo

    if idservico and vlrunitario is not None:
        gr_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            gr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_gr_data_into_db_manual(gr_data)

        else:
            profundidade_solo = max(float(rz_solo), float(
                profcobrancamin_tr) - float(rz_rocha))
            v_produzido = round(float(profundidade_solo) * float(vlrunitario) +
                                float(rz_rocha) * float(vlrunitariorocha), 2)
            gr_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_gr_data_into_db_manual(gr_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado.")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def verify_and_insert_hc_data(ordered_hc_data, idproposta):
    v_produzido = 0
    sigla = ordered_hc_data["sigla"]
    diametro = float(ordered_hc_data["diametro"])  # Convert diametro to float

    query = """
        SELECT idtiposervico
        FROM servicostipos
        WHERE sigla = %s
    """
    cursor = mydb.cursor()
    cursor.execute(query, (sigla,))
    result = cursor.fetchone()
    idtiposervico = result[0]
    print(f"idtiposervico: {idtiposervico}")

    query = """
        SELECT SUM(CASE WHEN servicos.flag_complementar = 'N' THEN propostasitens.qtd ELSE 0 END)
        FROM propostasitens
        JOIN servicos ON propostasitens.idservico = servicos.idservico
        WHERE propostasitens.idproposta = %s
    """
    cursor.execute(query, (idproposta,))
    total_qtd = cursor.fetchone()[0]
    print(f"total_qtd: {total_qtd}")

    query = """
        SELECT servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_hc, propostas.valorpctefechado
        FROM servicos
        JOIN propostasitens ON servicos.idservico = propostasitens.idservico
        JOIN obras ON propostasitens.idproposta = obras.idproposta
        JOIN propostas ON obras.idproposta = propostas.idproposta
        WHERE servicos.idtiposervico = %s AND propostasitens.idproposta = %s AND servicos.diametro = %s
        GROUP BY servicos.idservico, servicos.diametro, propostasitens.vlrunitario, obras.profcobrancamin_hc, propostas.valorpctefechado
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idtiposervico, idproposta, diametro))
    results = cursor.fetchall()

    idservico = None
    vlrunitario = None
    qtd = None
    for result in results:
        print(f"Current result: {result}")
        if float(result[1]) == diametro:  
            idservico = result[0]
            vlrunitario = result[2]
            profcobrancamin_hc = result[3]
            valorpctefechado = result[4]
            break

    print(f"idservico: {idservico}, vlrunitario: {vlrunitario}")

    profundidade = ordered_hc_data["profundidade"]

    concretagem_prevista = round(
        3.141592653589793 * ((diametro / 200) ** 2) * profundidade, 2)
    ordered_hc_data["concretagem_prevista"] = concretagem_prevista

    if idservico and vlrunitario is not None:
        ordered_hc_data["idservico"] = idservico

        if valorpctefechado > 0:
            v_produzido = round(float(valorpctefechado) / float(total_qtd), 2)
            ordered_hc_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_hc_data_into_db(ordered_hc_data)

        else:
            profundidade_solo = max(
                float(profundidade), float(profcobrancamin_hc))
            v_produzido = round(float(profundidade_solo) * float(vlrunitario))
            ordered_hc_data["v_produzido"] = v_produzido
            print(f"Calculated v_produzido: {v_produzido}")
            insert_hc_data_into_db(ordered_hc_data)

        return v_produzido

    else:
        warning_layout = [
            [sg.Text("O Diametro nao foi encontrado")],
            [sg.Button("OK")]
        ]
        warning_window = sg.Window("Warning", warning_layout)

        while True:
            event, values = warning_window.read()
            if event == sg.WIN_CLOSED or event == "OK":
                break

        warning_window.close()


def check_v_produzido_sum_rz(idproposta):
    cursor = mydb.cursor()
    query7 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query7, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_rz_manual(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_rt(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_rt_manual(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_pd(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_pd_manual(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def check_v_produzido_sum_hc(idproposta):
    cursor = mydb.cursor()
    query8 = """
        SELECT valorfaturminimo
        FROM propostas
        WHERE idproposta = %s
    """
    cursor.execute(query8, (idproposta,))
    result = cursor.fetchone()
    valorfaturminimo = result[0]
    return valorfaturminimo


def process_rz_service(ws, idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):

    column_mapping = {
        'Data': 'data_lcto',
        'Estaca': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Hora Início': 'hora_inicio',
        'Hora Fim': 'hora_termino',
        'Prof. Solo (m)': 'rz_solo',
        'Prof. Alt. (m)': 'rz_alteracao',
        'Prof. Rocha (m)': 'rz_rocha',
        'Inje. Hora Início': 'concretagem_inicio',
        'Inje. Hora Fim': 'concretagem_termino',
        'Scs cimento / Argamassa': 'rz_sacos_cimento',
        'Areia': 'rz_areia',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        rz_data_excel = dict(zip(column_mapping.keys(), row))

        rz_data = {column_mapping[key]: value for key,
                   value in rz_data_excel.items()}

        rz_data['idpatrimonio'] = idpatrimonio
        rz_data['idobra'] = idobra
        rz_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino', 'rz_solo', 'rz_alteracao',
                        'rz_rocha', 'concretagem_inicio', 'concretagem_termino', 'rz_sacos_cimento', 'rz_areia', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_rz_data = {key: rz_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_rz_data(ordered_rz_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_rz_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    query = """
        SELECT idservico
        FROM propostasitens
        WHERE idproposta = %s AND idservico = 508
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idproposta,))
    result = cursor.fetchone()

    if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 508:
        print("Faturamento Descoberto")
        valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
        rz_solo = valor_faltante / Decimal(50.0)
        new_row_data = {
            'idpatrimonio': idpatrimonio,
            'idobra': idobra,
            'sigla': 'RZ',
            'data_lcto': ordered_rz_data['data_lcto'],
            'estacaid': 'FMD',
            'diametro': 99.0,
            'secaoprojeto': 99.0,
            'hora_inicio': 0.0,
            'hora_termino': 0.0,
            'rz_solo': rz_solo,
            'rz_alteracao': 0.0,
            'rz_rocha': 0.0,
            'profundidade': 0.0,
            'concretagem_inicio': None,
            'concretagem_termino': None,
            'rz_sacos_cimento': 0.0,
            'rz_areia': 0.0,
            'observacoes': None,
            'flag_aprovada': 'S',
            'idservico': 508,
            'v_produzido': valor_faltante
        }

        insert_into_database_rz(new_row_data)

    return v_produzido_sum


def process_rz_service_manual(idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Estaca', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Hora Início', size=(20, 1)), sg.Input(key='hora_inicio')],
        [sg.Text('Hora Fim', size=(20, 1)), sg.Input(key='hora_termino')],
        [sg.Text('Prof. Solo (m)', size=(20, 1)), sg.Input(key='rz_solo')],
        [sg.Text('Prof. Alt. (m)', size=(20, 1)),
         sg.Input(key='rz_alteracao')],
        [sg.Text('Prof. Rocha (m)', size=(20, 1)), sg.Input(key='rz_rocha')],
        [sg.Text('Inje. Hora Início', size=(20, 1)),
         sg.Input(key='concretagem_inicio')],
        [sg.Text('Inje. Hora Fim', size=(20, 1)),
         sg.Input(key='concretagem_termino')],
        [sg.Text('Scs cimento / Argamassa', size=(20, 1)),
         sg.Input(key='rz_sacos_cimento')],
        [sg.Text('Areia', size=(20, 1)), sg.Input(key='rz_areia')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            rz_data = dict(values)

            for time_field in ['hora_inicio', 'hora_termino', 'concretagem_inicio', 'concretagem_termino']:
                if time_field not in rz_data or not rz_data[time_field]:
                    rz_data[time_field] = '00:00'

            for numeric_field in ['rz_sacos_cimento', 'rz_areia']:
                if numeric_field not in rz_data or not rz_data[numeric_field]:
                    rz_data[numeric_field] = '0.00'

            rz_data['idpatrimonio'] = idpatrimonio
            rz_data['idobra'] = idobra
            rz_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino', 'rz_solo', 'rz_alteracao',
                            'rz_rocha', 'concretagem_inicio', 'concretagem_termino', 'rz_sacos_cimento', 'rz_areia', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            rz_data = {
                k: rz_data[k] if k in rz_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_rz_data_manual(rz_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(rz_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')
            break  

        query = """
            SELECT idservico
            FROM propostasitens
            WHERE idproposta = %s AND idservico = 508
        """
        cursor = mydb.cursor()
        cursor.execute(query, (idproposta,))
        result = cursor.fetchone()

        if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 550:
            print("Faturamento Descoberto")
            valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
            rz_solo = valor_faltante / Decimal(50.0)
            new_row_data = {
                'idpatrimonio': idpatrimonio,
                'idobra': idobra,
                'sigla': 'RZ',
                'data_lcto': rz_data['data_lcto'],
                'estacaid': 'FMD',
                'diametro': 99.0,
                'secaoprojeto': 99.0,
                'hora_inicio': 0.0,
                'hora_termino': 0.0,
                'rz_solo': rz_solo,
                'rz_alteracao': 0.0,
                'rz_rocha': 0.0,
                'profundidade': 0.0,
                'concretagem_inicio': None,
                'concretagem_termino': None,
                'rz_sacos_cimento': 0.0,
                'rz_areia': 0.0,
                'observacoes': None,
                'flag_aprovada': 'S',
                'idservico': 508,
                'v_produzido': valor_faltante
            }

            insert_into_database_rz_manual(new_row_data)

    window.close()
    return v_produzido_sum


def process_rt_service(ws, idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):

    column_mapping = {
        'Data': 'data_lcto',
        'Estaca': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Prof. Escavada (m)': 'profundidade',
        'Prof. Útil (m)': 'concretagem_altura',
        'Seca / Submersa': 'Seca / Submersa',
        'Alargamento Seção (cm)': 'diametroalargamento',
        'Alargamento Prof. (m)': 'profundidadealargamento',
        'Camisa Seção (cm)': 'es_diametro',
        'Camisa Prof. (m)': 'es_profundidade',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        rt_data_excel = dict(zip(column_mapping.keys(), row))

        rt_data = {column_mapping[key]: value for key,
                   value in rt_data_excel.items()}

        rt_data['idpatrimonio'] = idpatrimonio
        rt_data['idobra'] = idobra
        rt_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade', 'concretagem_altura', 'Seca / Submersa',
                        'diametroalargamento', 'profundidadealargamento', 'es_diametro', 'es_profundidade', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_rt_data = {key: rt_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_rt_data(ordered_rt_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_rt_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    query = """
        SELECT idservico
        FROM propostasitens
        WHERE idproposta = %s AND idservico = 550
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idproposta,))
    result = cursor.fetchone()

    if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 550:
        print("Faturamento Descoberto")
        valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
        profundidade = valor_faltante / Decimal(50.0)
        new_row_data = {
            'idpatrimonio': idpatrimonio,
            'idobra': idobra,
            'sigla': 'RT',
            'data_lcto': ordered_rt_data['data_lcto'],
            'estacaid': 'FMD',
            'diametro': 99.0,
            'secaoprojeto': 99.0,
            'profundidade': profundidade,
            'concretagem_prevista': 0.0,
            'concretagem_realizada': 0.0,
            'concretagem_altura': 0.0,
            'es_flag_seca': 'N',
            'es_flag_submersa': 'N',
            'diametroalargamento': 0.0,
            'profundidadealargamento': 0.0,
            'es_diametro': 0.0,
            'es_profundidade': 0.0,
            'observacies': None,
            'flag_aprovada': 'S',
            'idservico': 550,
            'v_produzido': valor_faltante
        }

        insert_into_database_rt(new_row_data)

    return v_produzido_sum


def process_rt_service_manual(idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Estaca', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Prof. Escavada (m)', size=(20, 1)),
         sg.Input(key='profundidade')],
        [sg.Text('Prof. Útil (m)', size=(20, 1)),
         sg.Input(key='concretagem_altura')],
        [sg.Text('Seca / Submersa', size=(20, 1)), sg.Combo(['Seca',
                                                             'Submersa'], key='Seca / Submersa', default_value='Seca')],
        [sg.Text('Alargamento Seção (cm)', size=(20, 1)),
         sg.Input(key='diametroalargamento')],
        [sg.Text('Alargamento Prof. (m)', size=(20, 1)),
         sg.Input(key='profundidadealargamento')],
        [sg.Text('Camisa Seção (cm)', size=(20, 1)),
         sg.Input(key='es_diametro')],
        [sg.Text('Camisa Prof. (m)', size=(20, 1)),
         sg.Input(key='es_profundidade')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            rt_data = dict(values)

            for numeric_field in ['concretagem_altura', 'diametroalargamento', 'profundidadealargamento', 'es_diametro', 'es_profundidade']:
                if numeric_field not in rt_data or not rt_data[numeric_field]:
                    rt_data[numeric_field] = '0.00'

            rt_data['idpatrimonio'] = idpatrimonio
            rt_data['idobra'] = idobra
            rt_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade', 'concretagem_altura', 'Seca / Submersa',
                            'diametroalargamento', 'profundidadealargamento', 'es_diametro', 'es_profundidade', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            rt_data = {
                k: rt_data[k] if k in rt_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_rt_data_manual(rt_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(rt_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break  

        query = """
            SELECT idservico
            FROM propostasitens
            WHERE idproposta = %s AND idservico = 589
        """
        cursor = mydb.cursor()
        cursor.execute(query, (idproposta,))
        result = cursor.fetchone()

        if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 550:
            print("Faturamento Descoberto")
            valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
            profundidade = valor_faltante / Decimal(50.0)
            new_row_data = {
                'idpatrimonio': idpatrimonio,
                'idobra': idobra,
                'sigla': 'RT',
                'data_lcto': rt_data['data_lcto'],
                'estacaid': 'FMD',
                'diametro': 99.0,
                'secaoprojeto': 99.0,
                'profundidade': profundidade,
                'concretagem_prevista': 0.0,
                'concretagem_realizada': 0.0,
                'concretagem_altura': 0.0,
                'es_flag_seca': 'N',
                'es_flag_submersa': 'N',
                'diametroalargamento': 0.0,
                'profundidadealargamento': 0.0,
                'es_diametro': 0.0,
                'es_profundidade': 0.0,
                'observacies': None,
                'flag_aprovada': 'S',
                'idservico': 550,
                'v_produzido': valor_faltante
            }

            insert_into_database_rt_manual(new_row_data)

    window.close()
    return v_produzido_sum


def process_mt_service(ws, idpatrimonio, idobra, selected_service, idproposta):

    column_mapping = {
        'Data': 'data_lcto',
        'Estaca': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Prof. Cravada (m)': 'profundidade',
        'Segmentos utilizados': 'elementos',
        'Emendas': 'emendas',
        'Talas': 'talas',
        'Cortes': 'cortes',
        'Nega (mm)': 'nega',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        mt_data_excel = dict(zip(column_mapping.keys(), row))

        mt_data = {column_mapping[key]: value for key,
                   value in mt_data_excel.items()}

        mt_data['idpatrimonio'] = idpatrimonio
        mt_data['idobra'] = idobra
        mt_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade',
                        'elementos', 'emendas', 'talas', 'cortes', 'nega', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_mt_data = {key: mt_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_mt_data(ordered_mt_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_mt_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    return v_produzido_sum


def process_mt_service_manual(idpatrimonio, idobra, selected_service, idproposta):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Estaca', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Prof. Cravada (m)', size=(20, 1)),
         sg.Input(key='profundidade')],
        [sg.Text('Segmentos utilizados', size=(20, 1)),
         sg.Input(key='elementos')],
        [sg.Text('Emendas', size=(20, 1)), sg.Input(key='emendas')],
        [sg.Text('Talas', size=(20, 1)), sg.Input(key='talas')],
        [sg.Text('Cortes', size=(20, 1)), sg.Input(key='cortes')],
        [sg.Text('Nega (mm)', size=(20, 1)), sg.Input(key='nega')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            mt_data = dict(values)

            for numeric_field in ['elementos', 'emendas', 'talas', 'cortes', 'nega']:
                if numeric_field not in mt_data or not mt_data[numeric_field]:
                    mt_data[numeric_field] = '0.00'

            mt_data['idpatrimonio'] = idpatrimonio
            mt_data['idobra'] = idobra
            mt_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade',
                            'elementos', 'emendas', 'talas', 'cortes', 'nega', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            mt_data = {
                k: mt_data[k] if k in mt_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_mt_data_manual(mt_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(mt_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break 

    window.close()
    return v_produzido_sum


def process_pm_service(ws, idpatrimonio, idobra, selected_service, idproposta):

    column_mapping = {
        'Data': 'data_lcto',
        'Estaca': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Prof. Cravada (m)': 'profundidade',
        'Segmentos utilizados': 'elementos',
        'Emendas': 'emendas',
        'Talas': 'talas',
        'Cortes': 'cortes',
        'Nega (mm)': 'nega',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        pm_data_excel = dict(zip(column_mapping.keys(), row))

        pm_data = {column_mapping[key]: value for key,
                   value in pm_data_excel.items()}

        pm_data['idpatrimonio'] = idpatrimonio
        pm_data['idobra'] = idobra
        pm_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade',
                        'elementos', 'emendas', 'talas', 'cortes', 'nega', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_pm_data = {key: pm_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_pm_data(ordered_pm_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_pm_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    return v_produzido_sum


def process_pm_service_manual(idpatrimonio, idobra, selected_service, idproposta):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Estaca', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Prof. Cravada (m)', size=(20, 1)),
         sg.Input(key='profundidade')],
        [sg.Text('Segmentos utilizados', size=(20, 1)),
         sg.Input(key='elementos')],
        [sg.Text('Emendas', size=(20, 1)), sg.Input(key='emendas')],
        [sg.Text('Talas', size=(20, 1)), sg.Input(key='talas')],
        [sg.Text('Cortes', size=(20, 1)), sg.Input(key='cortes')],
        [sg.Text('Nega (mm)', size=(20, 1)), sg.Input(key='nega')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            pm_data = dict(values)

            for numeric_field in ['elementos', 'emendas', 'talas', 'cortes', 'nega']:
                if numeric_field not in pm_data or not pm_data[numeric_field]:
                    pm_data[numeric_field] = '0.00'

            pm_data['idpatrimonio'] = idpatrimonio
            pm_data['idobra'] = idobra
            pm_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'profundidade',
                            'elementos', 'emendas', 'talas', 'cortes', 'nega', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            pm_data = {
                k: pm_data[k] if k in pm_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_pm_data_manual(pm_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(pm_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break  

    window.close()
    return v_produzido_sum


def process_pd_service(ws, idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):

    column_mapping = {
        'Data': 'data_lcto',
        'Lamela': 'estacaid',
        'Espessura Executada (cm)': 'diametro',
        'Espessura Projeto (cm)': 'secaoprojeto',
        'Hora Início': 'hora_inicio',
        'Hora Fim': 'hora_termino',
        'Largura': 'largura',
        'Prof. (m)': 'profundidade',
        'Conc Hora Início': 'concretagem_inicio',
        'Conc Hora Fim': 'concretagem_termino',
        'Vol. (m³) Realizado': 'concretagem_realizada',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        print(f"Raw row data from Excel: {row}")

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        pd_data_excel = dict(zip(column_mapping.keys(), row))

        pd_data = {column_mapping[key]: value for key,
                   value in pd_data_excel.items()}

        pd_data['idpatrimonio'] = idpatrimonio
        pd_data['idobra'] = idobra
        pd_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino', 'largura', 'profundidade', 'area',
                        'concretagem_inicio', 'concretagem_termino', 'concretagem_realizada', 'concretagem_prevista', 'pd_flag_submersa', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_pd_data = {key: pd_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_pd_data(ordered_pd_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_pd_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    query = """
        SELECT idservico
        FROM propostasitens
        WHERE idproposta = %s AND idservico = 589
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idproposta,))
    result = cursor.fetchone()

    if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 589:
        print("Faturamento Descoberto")
        valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
        profundidade = valor_faltante / Decimal(50.0)
        new_row_data = {
            'idpatrimonio': idpatrimonio,
            'idobra': idobra,
            'sigla': 'PD',
            'data_lcto': ordered_pd_data['data_lcto'],
            'estacaid': 'FMD',
            'diametro': 99.0,
            'secaoprojeto': 99.0,
            'hora_inicio': 0.0,
            'hora_termino': 0.0,
            'largura': 0.0,
            'profundidade': profundidade,
            'area': "",
            'concretagem_inicio': 0.0,
            'concretagem_termino': 0.0,
            'concretagem_realizada': 0.0,
            'concretagem_prevista': 0.0,
            'pd_flag_submersa': 'S',
            'observacoes': None,
            'flag_aprovada': 'S',
            'idservico': 589,
            'v_produzido': valor_faltante
        }

        insert_into_database_pd(new_row_data)

    return v_produzido_sum


def process_pd_service_manual(idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Lamela', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Espessura Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Espessura Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Hora Início', size=(20, 1)), sg.Input(key='hora_inicio')],
        [sg.Text('Hora Fim', size=(20, 1)), sg.Input(key='hora_termino')],
        [sg.Text('Largura', size=(20, 1)), sg.Input(key='largura')],
        [sg.Text('Prof. (m)', size=(20, 1)), sg.Input(key='profundidade')],
        [sg.Text('Conc Hora Início', size=(20, 1)),
         sg.Input(key='concretagem_inicio')],
        [sg.Text('Conc Hora Fim', size=(20, 1)),
         sg.Input(key='concretagem_termino')],
        [sg.Text('Vol. (m³) Realizado', size=(20, 1)),
         sg.Input(key='concretagem_realizada')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            pd_data = dict(values)

            for numeric_field in ['concretagem_realizada']:
                if numeric_field not in pd_data or not pd_data[numeric_field]:
                    pd_data[numeric_field] = '0.00'

            for time_field in ['hora_inicio', 'hora_termino', 'concretagem_inicio', 'concretagem_termino']:
                if time_field not in pd_data or not pd_data[time_field]:
                    pd_data[time_field] = '00:00'

            pd_data['idpatrimonio'] = idpatrimonio
            pd_data['idobra'] = idobra
            pd_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino', 'largura', 'profundidade', 'area',
                            'concretagem_inicio', 'concretagem_termino', 'concretagem_realizada', 'concretagem_prevista', 'pd_flag_submersa', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            pd_data = {
                k: pd_data[k] if k in pd_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_pd_data_manual(pd_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(pd_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break 

        query = """
            SELECT idservico
            FROM propostasitens
            WHERE idproposta = %s AND idservico = 589
        """
        cursor = mydb.cursor()
        cursor.execute(query, (idproposta,))
        result = cursor.fetchone()

        if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 589:
            print("Faturamento Descoberto")
            valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
            profundidade = valor_faltante / Decimal(50.0)
            new_row_data = {
                'idpatrimonio': idpatrimonio,
                'idobra': idobra,
                'sigla': 'PD',
                'data_lcto': pd_data['data_lcto'],
                'estacaid': 'FMD',
                'diametro': 99.0,
                'secaoprojeto': 99.0,
                'hora_inicio': 0.0,
                'hora_termino': 0.0,
                'largura': 0.0,
                'profundidade': profundidade,
                'area': "",
                'concretagem_inicio': 0.0,
                'concretagem_termino': 0.0,
                'concretagem_realizada': 0.0,
                'concretagem_prevista': 0.0,
                'pd_flag_submersa': 'S',
                'observacoes': None,
                'flag_aprovada': 'S',
                'idservico': 589,
                'v_produzido': valor_faltante
            }

            insert_into_database_pd_manual(new_row_data)

    window.close()
    return v_produzido_sum


def process_tr_service(ws, idpatrimonio, idobra, selected_service, idproposta):

    column_mapping = {
        'Data': 'data_lcto',
        'Tirante': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Hora Início': 'hora_inicio',
        'Hora Fim': 'hora_termino',
        'Prof. (m)': 'profundidade',
        'Data Injeção': 'tr_datainjecao',
        'Data Projeção': 'tr_dataprojecao',
        'Prof. Rocha (m)': 'rz_rocha',
        'Scs cimento': 'rz_sacos_cimento',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        tr_data_excel = dict(zip(column_mapping.keys(), row))

        tr_data = {column_mapping[key]: value for key,
                   value in tr_data_excel.items()}

        if tr_data['tr_datainjecao'] == 0.00:
            tr_data['tr_datainjecao'] = None
        if tr_data['tr_dataprojecao'] == 0.00:
            tr_data['tr_dataprojecao'] = None

        tr_data['idpatrimonio'] = idpatrimonio
        tr_data['idobra'] = idobra
        tr_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino',
                        'profundidade', 'tr_datainjecao', 'tr_dataprojecao', 'rz_rocha', 'rz_sacos_cimento', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_tr_data = {key: tr_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_tr_data(ordered_tr_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_tr_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    return v_produzido_sum


def process_tr_service_manual(idpatrimonio, idobra, selected_service, idproposta):
    v_produzido_sum = 0 

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Tirante', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Hora Início', size=(20, 1)), sg.Input(key='hora_inicio')],
        [sg.Text('Hora Fim', size=(20, 1)), sg.Input(key='hora_termino')],
        [sg.Text('Prof. (m)', size=(20, 1)), sg.Input(key='profundidade')],
        [sg.Text('Data Injeção', size=(20, 1)), sg.Input(key="tr_datainjecao", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="tr_datainjecao", format="%Y-%m-%d")],
        [sg.Text('Data Projeção', size=(20, 1)), sg.Input(key="tr_dataprojecao", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="tr_dataprojecao", format="%Y-%m-%d")],
        [sg.Text('Prof. Rocha (m)', size=(20, 1)), sg.Input(key='rz_rocha')],
        [sg.Text('Scs cimento', size=(20, 1)),
         sg.Input(key='rz_sacos_cimento')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            tr_data = dict(values)

            for date_field in ['data_lcto', 'tr_datainjecao', 'tr_dataprojecao']:
                if not tr_data[date_field]:
                    tr_data[date_field] = None

            for numeric_field in ['rz_rocha', 'rz_sacos_cimento']:
                if numeric_field not in tr_data or not tr_data[numeric_field]:
                    tr_data[numeric_field] = '0.00'

            for time_field in ['hora_inicio', 'hora_termino']:
                if time_field not in tr_data or not tr_data[time_field]:
                    tr_data[time_field] = '00:00'

            tr_data['idpatrimonio'] = idpatrimonio
            tr_data['idobra'] = idobra
            tr_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino',
                            'profundidade', 'tr_datainjecao', 'tr_dataprojecao', 'rz_rocha', 'rz_sacos_cimento', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            tr_data = {
                k: tr_data[k] if k in tr_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_tr_data_manual(tr_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(tr_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break 

    window.close()
    return v_produzido_sum


def process_gr_service(ws, idpatrimonio, idobra, selected_service, idproposta):

    column_mapping = {
        'Data': 'data_lcto',
        'Grampo': 'estacaid',
        'Seção Executada (cm)': 'diametro',
        'Seção Projeto (cm)': 'secaoprojeto',
        'Hora Início': 'hora_inicio',
        'Hora Fim': 'hora_termino',
        'Prof. (m)': 'profundidade',
        'Data Injeção': 'tr_datainjecao',
        'Data Projeção': 'tr_dataprojecao',
        'Prof. Rocha (m)': 'rz_rocha',
        'Scs cimento': 'rz_sacos_cimento',
        'Observação': 'observacoes',
        'Aprovada (S/N)': 'flag_aprovada'
    }

    v_produzido_sum = 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):

        if all(cell_value is None for cell_value in row):
            continue

        row = [cell_value if cell_value is not None else 0.00 for cell_value in row]

        gr_data_excel = dict(zip(column_mapping.keys(), row))

        gr_data = {column_mapping[key]: value for key,
                   value in gr_data_excel.items()}

        gr_data['idpatrimonio'] = idpatrimonio
        gr_data['idobra'] = idobra
        gr_data['sigla'] = selected_service

        if gr_data['tr_datainjecao'] == 0.00:
            gr_data['tr_datainjecao'] = None
        if gr_data['tr_dataprojecao'] == 0.00:
            gr_data['tr_dataprojecao'] = None

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino',
                        'profundidade', 'tr_datainjecao', 'tr_dataprojecao', 'rz_rocha', 'rz_sacos_cimento', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_gr_data = {key: gr_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_gr_data(ordered_gr_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_gr_data)

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    return v_produzido_sum


def process_gr_service_manual(idpatrimonio, idobra, selected_service, idproposta):
    v_produzido_sum = 0  

    layout = [
        [sg.Text("Data de Lançamento", size=(20, 1)), sg.Input(key="data_lcto", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="data_lcto", format="%Y-%m-%d")],
        [sg.Text('Grampo', size=(20, 1)), sg.Input(key='estacaid')],
        [sg.Text('Seção Executada (cm)', size=(20, 1)),
         sg.Input(key='diametro')],
        [sg.Text('Seção Projeto (cm)', size=(20, 1)),
         sg.Input(key='secaoprojeto')],
        [sg.Text('Hora Início', size=(20, 1)), sg.Input(key='hora_inicio')],
        [sg.Text('Hora Fim', size=(20, 1)), sg.Input(key='hora_termino')],
        [sg.Text('Prof. (m)', size=(20, 1)), sg.Input(key='profundidade')],
        [sg.Text('Data Injeção', size=(20, 1)), sg.Input(key="tr_datainjecao", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="tr_datainjecao", format="%Y-%m-%d")],
        [sg.Text('Data Projeção', size=(20, 1)), sg.Input(key="tr_dataprojecao", do_not_clear=True),
         sg.CalendarButton("Selecione a Data", target="tr_dataprojecao", format="%Y-%m-%d")],
        [sg.Text('Prof. Rocha (m)', size=(20, 1)), sg.Input(key='rz_rocha')],
        [sg.Text('Scs cimento', size=(20, 1)),
         sg.Input(key='rz_sacos_cimento')],
        [sg.Text('Observação', size=(20, 1)), sg.Input(key='observacoes')],
        [sg.Text('Aprovada (S/N)', size=(20, 1)),
         sg.Combo(['S', 'N'], key='flag_aprovada', default_value='S')],
        [sg.Button('Enviar'), sg.Button('Voltar')]
    ]

    window = sg.Window('Tela de Inserção', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Voltar':
            break
        if event == 'Enviar':
            gr_data = dict(values)

            for date_field in ['data_lcto', 'tr_datainjecao', 'tr_dataprojecao']:
                if not gr_data[date_field]:
                    gr_data[date_field] = None

            for numeric_field in ['rz_rocha', 'rz_sacos_cimento']:
                if numeric_field not in gr_data or not gr_data[numeric_field]:
                    gr_data[numeric_field] = '0.00'

            for time_field in ['hora_inicio', 'hora_termino']:
                if time_field not in gr_data or not gr_data[time_field]:
                    gr_data[time_field] = '00:00'

            gr_data['idpatrimonio'] = idpatrimonio
            gr_data['idobra'] = idobra
            gr_data['sigla'] = selected_service

            ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino',
                            'profundidade', 'tr_datainjecao', 'tr_dataprojecao', 'rz_rocha', 'rz_sacos_cimento', 'observacoes', 'flag_aprovada', 'idservico', 'v_produzido']
            gr_data = {
                k: gr_data[k] if k in gr_data else None for k in ordered_keys}

            v_produzido = verify_and_insert_gr_data_manual(gr_data, idproposta)
            if v_produzido is not None:
                v_produzido_sum += v_produzido

            print(gr_data)

            print(
                f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

            sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

            for key in values.keys():
                window[key].update('')

            break 

    window.close()
    return v_produzido_sum


def process_hc_service(cursor, idpatrimonio, idobra, selected_service, idproposta, valorfaturminimo):

    column_mapping = {
        'estNumero': 'estacaid',
        'estDiametro': 'diametro',
        'estInicioP': 'hora_inicio',
        'estFimP': 'hora_termino',
        'estInicioC': 'concretagem_inicio',
        'estFimC': 'concretagem_termino',
        'estComprimento': 'profundidade',
        'estVolBetao': 'concretagem_realizada',
        'estSuperConsumo': 'sobreconsumo',
    }

    v_produzido_sum = 0.0

    mdb_columns = [
        'estNumero', 'estDiametro', 'estInicioP', 'estFimP',
        'estInicioC', 'estFimC', 'estComprimento', 'estVolBetao', 'estSuperConsumo',
    ]

    for row in cursor.fetchall():
        hc_data = {}
        for original_col, new_col in column_mapping.items():
            value = row[mdb_columns.index(original_col)]

            if new_col == 'hora_inicio':
                if isinstance(value, str):
                    dt_object = datetime.datetime.strptime(
                        value, '%Y-%m-%d %H:%M:%S')
                else:
                    dt_object = value
                hc_data['data_lcto'] = dt_object.date()
                value = dt_object.time()
            elif new_col == 'diametro':
                value = round(value * 100, 2)
                hc_data['secaoprojeto'] = value
            elif new_col in ['hora_termino', 'concretagem_inicio', 'concretagem_termino']:
                if isinstance(value, str):
                    value = datetime.datetime.strptime(
                        value, '%Y-%m-%d %H:%M:%S').time()
                else:
                    value = value.time()
            elif new_col in ['profundidade', 'concretagem_realizada', 'sobreconsumo']:
                value = round(value, 2)

            hc_data[new_col] = value

        hc_data['rz_solo'] = hc_data['profundidade']

        hc_data['flag_aprovada'] = 'S'

        hc_data['idpatrimonio'] = idpatrimonio
        hc_data['idobra'] = idobra
        hc_data['sigla'] = selected_service

        ordered_keys = ['idpatrimonio', 'idobra', 'sigla', 'data_lcto', 'estacaid', 'diametro', 'secaoprojeto', 'hora_inicio', 'hora_termino', 'concretagem_inicio',
                        'concretagem_termino', 'profundidade', 'rz_solo', 'concretagem_realizada', 'sobreconsumo', 'flag_aprovada', 'idservico', 'v_produzido']
        ordered_hc_data = {key: hc_data.get(key, None) for key in ordered_keys}

        v_produzido = verify_and_insert_hc_data(ordered_hc_data, idproposta)
        if v_produzido is not None:
            v_produzido_sum += v_produzido

        print(ordered_hc_data)
        print()  # Debug - Print an empty line to separate rows

    print(f"Sum of v_produzido for all lines inserted: {v_produzido_sum}")

    sg.popup("Dados inseridos com sucesso :).", title="Sucesso")

    query = """
        SELECT idservico
        FROM propostasitens
        WHERE idproposta = %s AND idservico = 496
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idproposta,))
    result = cursor.fetchone()

    if v_produzido_sum < valorfaturminimo and result is not None and result[0] == 496:
        print("Faturamento Descoberto")
        valor_faltante = valorfaturminimo - Decimal(v_produzido_sum)
        profundidade = valor_faltante / Decimal(50.0)
        new_row_data = {
            'idpatrimonio': idpatrimonio,
            'idobra': idobra,
            'sigla': 'HC',
            'data_lcto': ordered_hc_data['data_lcto'],
            'estacaid': 'FMD',
            'diametro': 99.0,
            'secaoprojeto': 99.0,
            'hora_inicio': None,
            'hora_termino': None,
            'concretagem_inicio': None,
            'concretagem_termino': None,
            'profundidade': profundidade,
            'rz_solo': profundidade,
            'concretagem_realizada': 0.0,
            'concretagem_prevista': 0.0,
            'sobreconsumo': 0.0,
            'observacoes': None,
            'flag_aprovada': 'S',
            'idservico': 496,
            'v_produzido': valor_faltante
        }

        insert_into_database_hc(new_row_data)

        return v_produzido_sum


def insert_into_database_rt(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        profundidade, concretagem_prevista, concretagem_realizada, concretagem_altura, es_flag_seca, es_flag_submersa, diametroalargamento, profundidadealargamento,
        es_diametro, es_profundidade, observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(profundidade)s, %(concretagem_prevista)s, %(concretagem_realizada)s, %(concretagem_altura)s, %(es_flag_seca)s, %(es_flag_submersa)s, %(diametroalargamento)s, %(profundidadealargamento)s,
        %(es_diametro)s, %(es_profundidade)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_rt_manual(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        profundidade, concretagem_prevista, concretagem_realizada, concretagem_altura, es_flag_seca, es_flag_submersa, diametroalargamento, profundidadealargamento,
        es_diametro, es_profundidade, observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(profundidade)s, %(concretagem_prevista)s, %(concretagem_realizada)s, %(concretagem_altura)s, %(es_flag_seca)s, %(es_flag_submersa)s, %(diametroalargamento)s, %(profundidadealargamento)s,
        %(es_diametro)s, %(es_profundidade)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_rz(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        hora_inicio, hora_termino, rz_solo, rz_alteracao, rz_rocha, profundidade,
        concretagem_inicio, concretagem_termino, rz_sacos_cimento, rz_areia,
        observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(hora_inicio)s, %(hora_termino)s, %(rz_solo)s, %(rz_alteracao)s, %(rz_rocha)s, %(profundidade)s,
        %(concretagem_inicio)s, %(concretagem_termino)s, %(rz_sacos_cimento)s, %(rz_areia)s,
        %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_rz_manual(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        hora_inicio, hora_termino, rz_solo, rz_alteracao, rz_rocha, profundidade,
        concretagem_inicio, concretagem_termino, rz_sacos_cimento, rz_areia,
        observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(hora_inicio)s, %(hora_termino)s, %(rz_solo)s, %(rz_alteracao)s, %(rz_rocha)s, %(profundidade)s,
        %(concretagem_inicio)s, %(concretagem_termino)s, %(rz_sacos_cimento)s, %(rz_areia)s,
        %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_pd(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        hora_inicio, hora_termino, profundidade, largura, area, concretagem_inicio,
        concretagem_termino, concretagem_realizada, pd_flag_submersa, observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(largura)s, %(area)s, %(concretagem_inicio)s,
        %(concretagem_termino)s, %(concretagem_realizada)s, %(pd_flag_submersa)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_pd_manual(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        hora_inicio, hora_termino, profundidade, largura, area, concretagem_inicio,
        concretagem_termino, concretagem_realizada, pd_flag_submersa, observacoes, flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(hora_inicio)s, %(hora_termino)s, %(profundidade)s, %(largura)s, %(area)s, %(concretagem_inicio)s,
        %(concretagem_termino)s, %(concretagem_realizada)s, %(pd_flag_submersa)s, %(observacoes)s, %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def insert_into_database_hc(row_data):
    cursor = mydb.cursor()
    sql = """
    INSERT INTO obras_producao (
        idpatrimonio, idobra, sigla, data_lcto, estacaid, diametro, secaoprojeto,
        hora_inicio, hora_termino, concretagem_inicio, concretagem_termino, profundidade, rz_solo, concretagem_realizada, concretagem_prevista, sobreconsumo,
        flag_aprovada, idservico, v_produzido
    ) VALUES (
        %(idpatrimonio)s, %(idobra)s, %(sigla)s, %(data_lcto)s, %(estacaid)s, %(diametro)s, %(secaoprojeto)s,
        %(hora_inicio)s, %(hora_termino)s, %(concretagem_inicio)s, %(concretagem_termino)s, %(profundidade)s, %(concretagem_realizada)s, %(concretagem_prevista)s, %(sobreconsumo)s,
        %(flag_aprovada)s, %(idservico)s, %(v_produzido)s
    )
    """
    try:
        cursor.execute(sql, row_data)
        mydb.commit()
    except Exception as e:
        sg.popup("Ocorreu um erro ao inserir dados no banco de dados. Por favor, verifique os dados e tente novamente. Detalhes do erro: " + str(e))


def show_machine_window(selected_machine, selected_service, proposta_codigo):
    if selected_service == "HC":
        file_type_text = "Selecione o MDB:"
        file_types = [("MDB Files", "*.mdb")]
    else:
        file_type_text = "Selecione o Excel:"
        file_types = [("Excel Files", "*.xlsx")]

    machine_layout = [
        [sg.Text(f"Maquina Selecionada: {selected_machine}")],
        [sg.Text(file_type_text), sg.Input(key="-FILE-"),
         sg.FileBrowse(file_types=file_types)],
        [sg.Button("Upload")]
    ]

    machine_window = sg.Window("Resultados", machine_layout)

    while True:
        event, values = machine_window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "Upload":
            try:
                file_path = values["-FILE-"]
                if file_path:
                    if file_path.endswith(".xlsx"):
                        if not check_excel_file(file_path):
                            sg.popup(
                                'Atenção', 'Caracteres proibidos detectado (Ç~)')
                            continue  
                    print("File uploaded!")
                    on_submit_proposta(
                        proposta_codigo, selected_service, selected_machine, file_path)
                else:
                    sg.popup("Please select a file before uploading.")
            except Exception as e:
                sg.popup(f"An error occurred: {str(e)}")


def check_excel_file(file_path):
    wb = load_workbook(filename=file_path, read_only=True)

    invalid_chars = 'ÇçÃã^`~'

    for sheet in wb:
        for row in sheet.iter_rows(values_only=True, min_row=2):
            for cell in row:
                if isinstance(cell, str) and any(char in cell for char in invalid_chars):
                    return False

    return True


def get_latest_obras_producao_entries(idobra):
    query = """
        SELECT data_lcto, estacaid, diametro, profundidade
        FROM obras_producao
        WHERE idobra = %s
        ORDER BY data_lcto DESC
        LIMIT 10
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idobra,))
    return cursor.fetchall()


def get_latest_horas_entries(idobra):
    query = """
        SELECT obras_ocorrencias.data_lcto, producaoocorrencias.descricao
        FROM obras_ocorrencias
        JOIN producaoocorrencias ON obras_ocorrencias.idocorrencia = producaoocorrencias.idocorrencia
        WHERE obras_ocorrencias.idobra = %s
        ORDER BY obras_ocorrencias.data_lcto DESC
        LIMIT 5
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idobra,))
    return cursor.fetchall()


def get_latest_equipes_entries(idobra):
    query = """
        SELECT obras_funcionarios.data_lcto, producaofuncionarios.nome
        FROM obras_funcionarios
        JOIN producaofuncionarios ON obras_funcionarios.idfuncionario = producaofuncionarios.idfuncionario
        WHERE obras_funcionarios.idobra = %s
        ORDER BY obras_funcionarios.data_lcto DESC
        LIMIT 5
    """
    cursor = mydb.cursor()
    cursor.execute(query, (idobra,))
    return cursor.fetchall()


def show_main_window():
    servico_options = ['RZ', 'RT', 'HC', 'MT', 'TR', 'PD', 'PM', 'GR']
    proposta_codigo_options = []

    layout = [
        [sg.Text("Selecione o servico:"), sg.Combo(servico_options, key="-SERVICO-", enable_events=True),
         sg.Text("Selecione o Codigo da Proposta:"),
         sg.Combo([f"{opt[0]} - {opt[1]}" for opt in proposta_codigo_options], key="-COMBO-", enable_events=True, disabled=True, size=(70, 10))],
        [sg.Button("Estacas Excel/MDB", key="-ESTACAS_EXCEL-", disabled=True),
         sg.Button("Estacas Manual", key="-ESTACAS_MANUAL-", disabled=True),
         sg.Button("Horas", key="-HORAS-", disabled=True),
         sg.Button("Equipes", key="-EQUIPES-", disabled=True),
         sg.Text("", size=(54, 1), key="-EXPAND-"),
         sg.Button("Logs", key="-LOGS-", disabled=True),
         sg.Button("Logout")]
    ]

    window = sg.Window("Sistema Provisório de Produção", layout, finalize=True)

    window['-EXPAND-'].expand(expand_x=True)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "Logout":
            try:
                os.remove(login_file_path)
            except Exception:
                pass
            window.close()
            show_login_window()
            break
        elif event == "-SERVICO-":
            proposta_codigo_options = get_proposta_codigo_options(
                values["-SERVICO-"])
            window["-COMBO-"].update(
                values=[f"{opt[0]} - {opt[1]}" for opt in proposta_codigo_options])

            window["-COMBO-"].update(disabled=False)

        elif event == "-COMBO-":
            if values["-SERVICO-"] and values["-COMBO-"]:
                window["-ESTACAS_EXCEL-"].update(disabled=False)
                window["-ESTACAS_MANUAL-"].update(disabled=False)
                window["-HORAS-"].update(disabled=False)
                window["-EQUIPES-"].update(disabled=False)
                window["-LOGS-"].update(disabled=False)

        elif event == "-ESTACAS_EXCEL-":
            selected_service = values["-SERVICO-"]
            selected_proposta_codigo_full = values["-COMBO-"]
            selected_proposta_codigo = selected_proposta_codigo_full.split(
                " - ")[0]  
            on_submit_proposta(selected_proposta_codigo, selected_service)

        elif event == "-ESTACAS_MANUAL-":
            selected_service = values["-SERVICO-"]
            selected_proposta_codigo_full = values["-COMBO-"]
            selected_proposta_codigo = selected_proposta_codigo_full.split(
                " - ")[0]  
            on_submit_proposta_manual(
                selected_proposta_codigo, selected_service)

        elif event == "-HORAS-":
            selected_service = values["-SERVICO-"]
            selected_proposta_codigo_full = values["-COMBO-"]
            selected_proposta_codigo = selected_proposta_codigo_full.split(
                " - ")[0]  
            on_submit_proposta_hours(selected_proposta_codigo)

        elif event == "-EQUIPES-":
            selected_service = values["-SERVICO-"]
            selected_proposta_codigo_full = values["-COMBO-"]
            selected_proposta_codigo = selected_proposta_codigo_full.split(
                " - ")[0]  
            on_submit_proposta_teams(selected_proposta_codigo)

        elif event == "-LOGS-":
            selected_service = values["-SERVICO-"]
            selected_proposta_codigo_full = values["-COMBO-"]
            selected_proposta_codigo = selected_proposta_codigo_full.split(
                " - ")[0]  

            idobra = get_idobra_for_proposta_codigo(selected_proposta_codigo)
            latest_estacas_entries = get_latest_obras_producao_entries(idobra)
            latest_horas_entries = get_latest_horas_entries(idobra)
            latest_equipes_entries = get_latest_equipes_entries(idobra)

            layout = [[sg.Text(f"Logs para {selected_proposta_codigo}")],
                      [sg.Text(" ")],
                      [sg.Text("Estacas")]]

            for entry in latest_estacas_entries:
                layout.append([sg.Text(
                    f"Data: {entry[0]}, Estaca: {entry[1]}, Diametro: {entry[2]}, Profundidade: {entry[3]}")])

            layout.append([sg.Text(" ")])

            layout.append([sg.Text("Horas")])

            for entry in latest_horas_entries:
                layout.append(
                    [sg.Text(f"Data: {entry[0]}, Ocorrencia: {entry[1]}")])

            layout.append([sg.Text(" ")])

            layout.append([sg.Text("Equipes")])

            for entry in latest_equipes_entries:
                layout.append([sg.Text(f"Data: {entry[0]}, Nome: {entry[1]}")])

            log_window = sg.Window("Logs", layout)

            while True:
                log_event, log_values = log_window.read()
                if log_event == sg.WIN_CLOSED:
                    log_window.close()
                    break

    window.close()

show_login_window()
